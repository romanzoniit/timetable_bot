import openpyxl
import parsetable
import enum
import re
import xlsx_env
import pandas as pd
book = openpyxl.open(parsetable.timetable, read_only=True)

sheet = book.active

def parse_day():
     dict_l=dict()
     for j in range(sheet.min_column-1, sheet.max_column):
          for i in range(sheet.min_row-1,sheet.max_row):
               if sheet[i][j].value != None and sheet[i][j].value in xlsx_env.dday.keys():
                    day = sheet[i][j].value
                    time_idx = i
                    dict_l[day] = []
                    while sheet[time_idx][j+1].value != xlsx_env.dtime[-1]:
                         if sheet[time_idx][j + 1].value in xlsx_env.dtime:
                              dict_l[day].append({str(sheet[time_idx][j+1].value).replace(' ', ''): time_idx})
                         time_idx += 1
                    dict_l[day].append({str(sheet[time_idx][j+1].value).replace(' ', ''): time_idx})
                    if sheet[time_idx][j + 1].value == xlsx_env.dtime[-1] and day == list(xlsx_env.dday.keys())[-1]:
                         print(dict_l)
                         return dict_l

#data = parse_day()
"""def parse_lesson_group():
     all_info = []
     for i in range(sheet.min_row-1,sheet.max_row):
          for j in range(sheet.min_column - 1, sheet.max_column):
               if sheet[i][j].value != None and sheet[i][j].value in xlsx_env.dgroup:
                    group_idx = j
                    group = sheet[i][j].value
                    for idx, elem in enumerate(data)
                         all_info.append((parse_day()))"""


def parse_group_idx():
     dict_l = dict()
     for j in range(sheet.min_column - 1, sheet.max_column):
          for i in range(sheet.min_row - 1, sheet.max_row):
               if sheet[i][j].value == list(xlsx_env.dday.keys())[0] or sheet[i][j].value == xlsx_env.dtime[0]:
                    break
               if sheet[i][j].value != None and sheet[i][j].value in xlsx_env.dgroup:
                    group_idx = j
                    while sheet[i][group_idx].value != xlsx_env.dgroup[-1]:
                         if sheet[i][group_idx].value in xlsx_env.dgroup:
                              dict_l[sheet[i][group_idx].value]=group_idx
                         group_idx += 1
                    dict_l[sheet[i][group_idx].value] = group_idx
                    return dict_l

def parse_time():
     time_list = []
     for row in range(33, 46, 2):
          time_list.append(sheet[row][1].value)
     return time_list

def parse_groups():
     groups = []
     for column in range(2,sheet.max_column):
          groups.append(sheet[32][column].value)
     return groups

if __name__ == "__main__":
     print(parse_group_idx())
     print(parse_day())